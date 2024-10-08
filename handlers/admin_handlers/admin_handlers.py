import os
import sqlite3

import openpyxl
from aiogram import types
from aiogram.filters import Command
from aiogram.filters import StateFilter
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import StatesGroup, State
from aiogram.types import FSInputFile
from aiogram.types import Message
from loguru import logger

from system.dispatcher import ADMIN_USER_ID
from system.dispatcher import bot
from system.dispatcher import dp
from system.dispatcher import router


@router.message(Command("help"))
async def help_handler(message: Message, state: FSMContext):
    """Админ панель"""
    if message.from_user.id == ADMIN_USER_ID:
        await message.answer("Команды админа:\n\n"

                             "<b>Редактирование текста:</b>\n"
                             "/edit_services_and_prices - редактирование: ⭐️ Услуги и цены\n"
                             "/edit - редактирование: пост приветствиt\n"
                             "/edit_self_purchase - редактирование: 🛍 Самовыкуп\n"
                             "/edit_product_search - редактирование: Подбор товара\n"
                             "/edit_search_in_china - редактирование: Поиск поставщика в Китае\n"
                             "/edit_warranty_service - редактирование: Выкуп товаров\n"
                             "/edit_delivery_in_china - редактирование: Доставка в Китае\n"
                             "/edit_payment_process - редактирование: Как совершается оплата?\n"
                             "/edit_order_form - редактирование: 🗒 Бланк заказа\n"
                             "/edit_payment_options - редактирование: Какие платежи меня ожидают?\n"
                             "/edit_reviews - редактирование: 💌 Отзывы\n"
                             "/edit_bag_tape - редактирование: Мешок + скотч\n"
                             "/edit_box_bag_tape - редактирование: Коробка + мешок + скотч\n"
                             "/edit_cardboard_corners_bag_tape - редактирование: Картонные уголки + мешок + скотч\n"
                             "/edit_pallet_crate - редактирование: Паллет в обрешетке\n"
                             "edit_pallet_with_solid_wooden_box - редактирование: Паллет с глухим деревянным коробом\n"
                             "/edit_types_packaging_handlers - редактирование: Назад к видам упаковки\n"
                             "/edit_wooden_sheathing_bag_tape - редактирование: Деревянная обрешетка + мешок + скотч\n"
                             "/edit_useful_information - редактирование: 📚 Полезная информация\n\n"

                             "<b>Получение данных:</b>\n"
                             "/get_a_list_of_users_registered_in_the_bot - Получение списка зарегистрированных пользователей\n"
                             "/get_users_who_launched_the_bot - Получение данных пользователей, запускающих бота\n\n"

                             "<b>Отправка сообщений:</b>\n"
                             "/send_an_image_to_bot_users - Отправка изображения через бота + текст\n"
                             "/send_a_message_to_bot_users - Отправка текста через бота\n\n"
                             
                             "<b>Замена изображения поста:</b>\n"
                             "/edit_photo - редактирование фото главного поста\n"
                             "/delivery_in_china_photo - ⭐️ Услуги и цены\n"
                             "/warranty_service_photo - Выкуп товаров\n"
                             "/product_search_photo - Подбор товара\n"
                             "/self_purchase_photo - 🛍 Самовыкуп\n\n"
                             "/start - начальное меню\n", parse_mode="HTML")
    else:
        await message.reply("У вас нет прав на выполнение этой команды.", parse_mode="HTML")


# Функция для создания файла Excel с данными заказов
def create_excel_file(orders):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    # Заголовки столбцов
    sheet['A1'] = 'ID аккаунта пользователя'
    sheet['B1'] = 'Имя'
    sheet['C1'] = 'Фамилия'
    sheet['D1'] = 'Город'
    sheet['E1'] = 'Номер телефона'
    sheet['F1'] = 'Дата регистрации'
    # Заполнение данными заказов
    for index, order in enumerate(orders, start=2):
        sheet.cell(row=index, column=1).value = order[0]  # ID аккаунта пользователя
        sheet.cell(row=index, column=2).value = order[1]  # Имя
        sheet.cell(row=index, column=3).value = order[2]  # Фамилия
        sheet.cell(row=index, column=4).value = order[3]  # Город
        sheet.cell(row=index, column=5).value = order[4]  # Номер телефона
        sheet.cell(row=index, column=6).value = order[5]  # Дата регистрации

    return workbook


@router.message(Command("get_a_list_of_users_registered_in_the_bot"))
async def export_data(message: types.Message, state: FSMContext):
    """Получение списка зарегистрированных пользователей"""
    await state.clear()  # Очищаем состояние
    try:
        if message.from_user.id not in [535185511, 535773227, 388207248]:
            await message.reply('У вас нет доступа к этой команде.', parse_mode="HTML")
            return
        # Подключение к базе данных SQLite
        conn = sqlite3.connect('your_database.db')
        cursor = conn.cursor()
        # Получение данных из базы данных
        cursor.execute("SELECT * FROM users")
        orders = cursor.fetchall()
        # Создание файла Excel
        workbook = create_excel_file(orders)
        filename = 'Зарегистрированные пользователи в боте.xlsx'
        workbook.save(filename)  # Сохранение файла
        text = ("Данные пользователей зарегистрированных в боте\n\n"
                "Для возврата в начальное меню нажми на /start или /help")
        file = FSInputFile(filename)
        await bot.send_document(message.from_user.id, document=file, caption=text, parse_mode="HTML")  # Отправка файла пользователю
        os.remove(filename)  # Удаление файла
    except Exception as e:
        logger.error(e)


def create_excel_file_start(orders):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    # Заголовки столбцов
    sheet['A1'] = 'ID аккаунта пользователя'
    sheet['B1'] = 'username'
    sheet['C1'] = 'Имя'
    sheet['D1'] = 'Фамилия'
    sheet['E1'] = 'Дата запуска бота'
    # Заполнение данными заказов
    for index, order in enumerate(orders, start=2):
        sheet.cell(row=index, column=1).value = order[0]  # ID аккаунта пользователя
        sheet.cell(row=index, column=2).value = order[1]  # username
        sheet.cell(row=index, column=3).value = order[2]  # Имя
        sheet.cell(row=index, column=4).value = order[3]  # Фамилия
        sheet.cell(row=index, column=5).value = order[4]  # Дата запуска бота

    return workbook


@router.message(Command("get_users_who_launched_the_bot"))
async def get_users_who_launched_the_bot(message: types.Message, state: FSMContext):
    """Получение данных пользователей, запускающих бота"""
    await state.clear()  # Очищаем состояние
    try:
        if message.from_user.id not in [535185511, 535773227, 388207248]:
            await message.reply('У вас нет доступа к этой команде.', parse_mode="HTML")
            return
        conn = sqlite3.connect('your_database.db')  # Подключение к базе данных SQLite
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM users_start")  # Получение данных из базы данных
        orders = cursor.fetchall()
        workbook = create_excel_file_start(orders)  # Создание файла Excel
        filename = 'Данные пользователей запустивших бота.xlsx'
        workbook.save(filename)  # Сохранение файла
        file = FSInputFile(filename)
        text = ("Данные пользователей зарегистрированных в боте\n\n"
                "Для возврата в начальное меню нажми на /start или /help")
        await bot.send_document(message.from_user.id, document=file, caption=text, parse_mode="HTML")  # Отправка файла пользователю
        os.remove(filename)  # Удаление файла
    except Exception as e:
        logger.error(e)


class MyStates(StatesGroup):
    waiting_for_message = State()
    waiting_for_image = State()
    waiting_for_caption = State()


@router.message(Command("send_an_image_to_bot_users"))
async def send_an_image_to_bot_users(message: types.Message, state: FSMContext):
    """Запрашивает изображение у администратора"""
    await state.clear()  # Очищаем состояние
    try:
        if message.from_user.id not in [535185511, 535773227, 388207248]:
            await message.reply('У вас нет доступа к этой команде.')
            return
        await bot.send_message(message.from_user.id, text="Загрузите изображение для рассылки:", parse_mode="HTML")
        await state.set_state(MyStates.waiting_for_image)
    except Exception as e:
        logger.error(e)


@router.message(StateFilter(MyStates.waiting_for_image))
async def process_send_image(message: types.Message, state: FSMContext):
    """
    Этот хендлер будет ждать загруженного изображения и переходить в состояние "ожидание подписи"
    """

    await state.update_data(photo=message.photo[-1].file_id)
    await bot.send_message(message.from_user.id, text="Введите подпись к изображению:", parse_mode="HTML")
    await state.set_state(MyStates.waiting_for_caption)


@router.message(StateFilter(MyStates.waiting_for_caption))
async def process_send_image_with_caption(message: types.Message, state: FSMContext):
    """
    Этот хендлер будет ждать введенной подписи и выполнять рассылку
    """
    state_data = await state.get_data()  # Получить данные о состоянии
    state_data['caption'] = message.text  # Сохраните заголовок в данных состояния
    photo = state_data.get('photo')  # Получите фотографию и подпись из государственных данных.
    caption = state_data.get('caption')
    user_ids = get_user_ids()  # Получаем список уникальных ID пользователей из базы данных
    if user_ids:
        for user_id in user_ids:  # Рассылка изображения с подписью всем пользователям из списка
            try:
                await bot.send_photo(user_id, photo, caption=caption)  # Отправляем изображение с подписью
            except Exception as e:
                print(f"Ошибка при отправке изображения с подписью пользователю {user_id}: {str(e)}")
    await message.answer("Изображение успешно разослано всем пользователям.", parse_mode="HTML")
    await state.clear()


@router.message(Command("send_a_message_to_bot_users"))
async def send_a_message_to_bot_users(message: types.Message, state: FSMContext):
    """Запрашивает текст сообщения у администратора"""
    await state.clear()
    try:
        if message.from_user.id not in [535185511, 535773227, 388207248]:
            await message.reply('У вас нет доступа к этой команде.')
            return
        await bot.send_message(message.from_user.id, text="Введите текст для рассылки:", parse_mode="HTML")
        await state.set_state(MyStates.waiting_for_message)
    except Exception as e:
        logger.error(e)


@router.message(StateFilter(MyStates.waiting_for_message))
async def process_send_message(message: types.Message, state: FSMContext):
    """
    Этот хендлер будет ждать введенного текста и выполнять рассылку
    """
    # Получаем список уникальных ID пользователей из базы данных
    message_text = message.text
    user_ids = get_user_ids()
    if user_ids:
        for user_id in user_ids:  # Рассылка сообщения всем пользователям из списка
            try:
                await bot.send_message(chat_id=user_id, text=message_text, parse_mode="HTML")
            except Exception as e:
                print(f"Ошибка при отправке сообщения пользователю {user_id}: {str(e)}")
    await message.answer("Сообщение успешно разослано всем пользователям.", parse_mode="HTML")
    await state.clear()


def get_user_ids():
    """Получаем уникальные ID пользователей из базы данных"""
    try:
        conn = sqlite3.connect('your_database.db')  # Замените 'your_database.db' на имя вашей базы данных
        cursor = conn.cursor()
        cursor.execute("SELECT DISTINCT user_id FROM users_start")
        user_ids = [row[0] for row in cursor.fetchall()]
        return user_ids
    except Exception as e:
        print(f"Ошибка при получении ID пользователей из базы данных: {str(e)}")
        return []


def register_handlers_admin():
    dp.message.register(help_handler)
    dp.message.register(export_data)
    dp.message.register(get_users_who_launched_the_bot)
    dp.message.register(send_a_message_to_bot_users)
    dp.message.register(send_an_image_to_bot_users)
